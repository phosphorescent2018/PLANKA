from flask import Flask, request, jsonify
import sqlite3
import json
import os
import re
from datetime import datetime

app = Flask(__name__)
DB_FILE = 'planka_events.db'

def init_db():
    """初始化数据库，包含自动升级(Migration)逻辑"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # 1. 确保基础表存在
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_type TEXT,
            item_name TEXT,
            board_name TEXT,
            user_name TEXT,
            raw_data TEXT,
            received_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 2. 检查并添加新字段 (Migration)
    # 获取当前所有字段名
    cursor.execute("PRAGMA table_info(events)")
    existing_columns = [col[1] for col in cursor.fetchall()]
    
    # 定义需要新增的字段
    new_columns = {
        'card_id': 'TEXT',
        'from_list': 'TEXT',
        'to_list': 'TEXT'
    }
    
    for col_name, col_type in new_columns.items():
        if col_name not in existing_columns:
            print(f"🔄 正在升级数据库: 添加字段 {col_name}...")
            cursor.execute(f"ALTER TABLE events ADD COLUMN {col_name} {col_type}")
    
    conn.commit()
    conn.close()
    print(f"✅ 数据库已就绪 (Pro v2.0): {os.path.abspath(DB_FILE)}")

@app.route('/webhook', methods=['POST'])
def handle_webhook():
    """接收并处理 Planka 的 Webhook 数据 (Markdown 解析版)"""
    try:
        data = request.json
        if not data:
            return jsonify({"status": "error", "message": "No JSON payload"}), 400

        # 初始化默认值
        event_type = 'Unknown'
        item_name = 'N/A'
        board_name = 'N/A'
        user_name = 'System'
        card_id = None
        from_list = None
        to_list = None
        
        # ========= 核心解析逻辑 =========
        # 场景 A: Apprise (UI配置 json://... 且格式为 Markdown)
        if 'message' in data:
            event_type = data.get('title', 'Notification')
            raw_message = data.get('message', '')
            
            # 1. 提取操作人 (通常是第一个词)
            user_name = raw_message.split(' ')[0]
            
            # 2. 提取看版名 (on 之后)
            # Markdown 格式通常是: ... on 看板名 (可能没加粗，也可能加粗)
            match_board = re.search(r' on (.*?)$', raw_message)
            if match_board:
                board_name = match_board.group(1).strip()

            # 3. 提取卡片名和 Card ID (这是 Markdown 模式的核心优势)
            # 格式: [CardName](http://.../cards/card-uuid)
            match_card = re.search(r'\[(.*?)\]\((.*?/cards/([a-zA-Z0-9-]+))\)', raw_message)
            if match_card:
                item_name = match_card.group(1) # 卡片名
                card_id = match_card.group(3)   # ID (URL的最后一部分)
            else:
                # 兼容如果不小心还是发了 Text 格式的情况
                match_text_card = re.search(r'\[(.*?)\]', raw_message)
                if match_text_card:
                    item_name = match_text_card.group(1)

            # 4. 提取流转列表 (From -> To)
            # 格式: from **ListA** to **ListB**
            match_move = re.search(r'from \*\*(.*?)\*\* to \*\*(.*?)\*\*', raw_message)
            if match_move:
                from_list = match_move.group(1)
                to_list = match_move.group(2)

        # 场景 B: 原生 Webhook (备用)
        elif 'event' in data:
            event_type = data.get('event', 'unknown')
            payload = data.get('data', {})
            item = payload.get('item', {})
            item_name = item.get('name', '')
            card_id = item.get('id', '')
            # ... 原生格式暂不深度展开，优先保障 Apprise Markdown

        # 写入数据库 (包含新字段)
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO events (event_type, item_name, board_name, user_name, card_id, from_list, to_list, raw_data)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (event_type, item_name, board_name, user_name, card_id, from_list, to_list, json.dumps(data, ensure_ascii=False)))
        conn.commit()
        conn.close()

        # ========= 企业微信推送 (带过滤) =========
        # 配置区
        WECOM_WEBHOOK = os.environ.get('WECOM_WEBHOOK', 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=04113332-f1c3-482b-9cd3-83ba7d5e99ef')
        ALLOWED_BOARDS = ['EP']  # 只推送这些看板
        ALLOWED_TYPES = ['Card Moved', 'Card Created']  # 只推送这些类型
        
        # 判断是否需要推送
        should_push = (board_name in ALLOWED_BOARDS) and (event_type in ALLOWED_TYPES)
        
        if should_push and WECOM_WEBHOOK:
            import requests
            
            # 翻译事件类型
            type_cn = {
                'Card Moved': '📋 卡片移动',
                'Card Created': '✨ 卡片创建',
                'New Comment': '💬 新评论',
                'You Were Added to Card': '👤 被添加到卡片'
            }.get(event_type, event_type)
            
            # 构建 Markdown 消息
            if event_type == 'Card Moved':
                content = f'''{type_cn}
> 卡片: <font color="info">{item_name}</font>
> 操作人: {user_name}
> 看板: {board_name}
> 流转: <font color="warning">{from_list} → {to_list}</font>'''
            else:
                content = f'''{type_cn}
> 卡片: <font color="info">{item_name}</font>
> 操作人: {user_name}
> 看板: {board_name}'''
            
            # 如果有卡片链接，添加链接
            raw = json.loads(json.dumps(data))
            if 'message' in raw:
                import re as regex
                link_match = regex.search(r'\((https?://[^)]+)\)', raw.get('message', ''))
                if link_match:
                    content += f'\n\n[🔗 点击查看卡片]({link_match.group(1)})'
            
            # 发送到企业微信
            try:
                resp = requests.post(WECOM_WEBHOOK, json={
                    "msgtype": "markdown",
                    "markdown": {"content": content}
                }, timeout=5)
                print(f"📤 企业微信推送: {resp.status_code}")
            except Exception as e:
                print(f"⚠️ 企业微信推送失败: {e}")

        # 打印详细日志
        print(f"\n[📝 新记录] {datetime.now().strftime('%H:%M:%S')}")
        print(f"类型: {event_type}")
        print(f"用户: {user_name} -> 看板: {board_name}")
        if card_id:
            print(f"卡片: {item_name} (ID: {card_id})")
        else:
            print(f"内容: {item_name}")
            
        if from_list and to_list:
            print(f"流转: {from_list} ➡️  {to_list}")
        
        if should_push:
            print(f"📤 已推送到企业微信")
        else:
            print(f"⏭️ 未推送 (不在白名单)")

        return jsonify({"status": "success"}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ 处理出错: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/events', methods=['GET'])
def list_events():
    """API: 查看最近数据 (含新字段)"""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM events ORDER BY received_at DESC LIMIT 10')
    rows = cursor.fetchall()
    conn.close()
    return jsonify([dict(row) for row in rows])

@app.route('/download', methods=['GET'])
def download_excel():
    """下载全部数据为 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file
    
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM events ORDER BY received_at DESC')
    rows = cursor.fetchall()
    conn.close()
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Planka Events"
    
    # 定义表头 (用户需要的字段 + 额外有用的字段)
    headers = [
        'ID', 
        '时间', 
        '类型', 
        '卡片名', 
        'Card ID', 
        '操作人', 
        '看板', 
        '流转前列表', 
        '流转后列表',
        '原始消息'  # 保留完整消息，方便排查
    ]
    
    # 写入表头并设置样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 写入数据
    for row_idx, row in enumerate(rows, 2):
        data = dict(row)
        # 提取原始消息中的 message 字段
        raw = json.loads(data.get('raw_data', '{}'))
        original_message = raw.get('message', '')
        
        ws.cell(row=row_idx, column=1, value=data.get('id'))
        ws.cell(row=row_idx, column=2, value=data.get('received_at'))
        ws.cell(row=row_idx, column=3, value=data.get('event_type'))
        ws.cell(row=row_idx, column=4, value=data.get('item_name'))
        ws.cell(row=row_idx, column=5, value=data.get('card_id'))
        ws.cell(row=row_idx, column=6, value=data.get('user_name'))
        ws.cell(row=row_idx, column=7, value=data.get('board_name'))
        ws.cell(row=row_idx, column=8, value=data.get('from_list'))
        ws.cell(row=row_idx, column=9, value=data.get('to_list'))
        ws.cell(row=row_idx, column=10, value=original_message)
    
    # 调整列宽
    column_widths = [6, 20, 15, 30, 25, 15, 15, 15, 15, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 保存到内存并返回
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"planka_events_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    init_db()
    # 允许通过环境变量设置端口，方便云端部署
    port = int(os.environ.get('PORT', 5000))
    
    print("------------------------------------------")
    print("📡 Planka 收集器 Pro v2.0 (Markdown版) 已启动")
    print("   请确保 Planka 通知格式已设为: Markdown")
    print(f"   监听端口: {port}")
    print("------------------------------------------")
    app.run(host='0.0.0.0', port=port)
